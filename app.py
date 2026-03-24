"""
Falafel Bribua - Daily Sales Reporting Dashboard
=================================================
Streamlit app that aggregates data from multiple sources (CSV, Excel, PDF)
into a unified daily report matching the master file format.
"""

import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
import os
import datetime
from pathlib import Path
from collections import defaultdict

# ──────────────────────────────────────────────
# Page config
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="פלאפל בריבוע - דוח יומי",
    page_icon="🧆",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────

# Branch name mapping: source name → master name
BRANCH_NAME_MAP = {
    # Non-Paz (CSV / Excel sources)
    "פלאפל בריבוע אשדוד": "אשדוד",
    "פלאפל בריבוע בית עובד": "בית עובד",
    "פלאפל בריבוע גבעת שמואל": "גבעת שמואל",
    "פלאפל בריבוע גן יבנה": "גן יבנה",
    "פלאפל בריבוע גני תקוה": "גני תקווה",
    "פלאפל בריבוע גני תקווה": "גני תקווה",
    "פלאפל בריבוע חולון": "חולון",
    "פלאפל בריבוע יבנה": "יבנה",
    "פלאפל בריבוע מרום נווה": "מרום נווה",
    "פלאפל בריבוע נתניה": "נתניה",
    "פלאפל בריבוע סניף המפעל": "סניף המפעל",
    "פלאפל בריבוע עין שמר": "עין שמר",
    "פלאפל בריבוע צורן": "צורן",
    "פלאפל בריבוע ראש העין": "ראש העין",
    "פלאפל בריבוע ראשל\"צ": "רוטשילד",
    'פלאפל בריבוע ראשל"צ': "רוטשילד",
    "פלאפל בריבוע רחובות": "רחובות",
    "פלאפל בריבוע רמת השרון": "רמת השרון",
    "פלאפל בריבוע שוהם": "שוהם",
    # Paz branches (PDF sources)
    "מתחם זך - ליאת חג'ג'": "רופין",
    "מתחם זך": "רופין",
    "ליאת חג'ג'": "רופין",
    "מתחם טופז -שחר בוני": "טופז",
    "מתחם טופז": "טופז",
    "שחר בוני": "טופז",
    "טופז": "טופז",
    "פלאפל - השקמה": "שקמה",
    "השקמה": "שקמה",
    "פלאפל - מנחם": "מנחם",
    "מנחם": "מנחם",
    "פלאפל אשכול": "אשכול",
    "אשכול": "אשכול",
    "פלאפל נחלת יהודה": "נחלת",
    "נחלת יהודה": "נחלת",
    "נחלת": "נחלת",
    "פלאפל שוהם": "פז שוהם",
    # In Paz PDF context, "פלאפל שוהם" = פז שוהם
    # but in non-Paz context "פלאפל בריבוע שוהם" = שוהם
}

# Paz branch names as they appear in the master file
PAZ_BRANCHES = {"רופין", "טופז", "מנחם", "פז שוהם", "אשכול", "נחלת", "שקמה"}

# Master column order
MASTER_COLUMNS = [
    "סניף",
    'מכר כולל מע"מ',
    "ממוצע עסקאות",
    "מס' עסקאות",
    "עסקה ראשונה",
    "עסקה אחרונה",
    "מנות בפיתה",
    "ארוחות בפיתה",
    "אחוז ארוחות מתוך מנות",
]

# Master branch order (as it appears in the original file)
MASTER_BRANCH_ORDER = [
    "אשדוד", "בית עובד", "גבעת שמואל", "גן יבנה", "גני תקווה",
    "חולון", "יבנה", "מרום נווה", "נתניה", "סניף המפעל",
    "עין שמר", "צורן", "ראש העין", "רוטשילד", "רחובות",
    "רמת השרון", "שוהם",
    "רופין", "טופז", "מנחם", "פז שוהם", "אשכול", "נחלת", "שקמה",
]


# ──────────────────────────────────────────────
# Helper: normalize branch name
# ──────────────────────────────────────────────
def normalize_branch(name: str, is_paz_context: bool = False) -> str:
    """Map a raw branch name to the master file's branch name."""
    if not name or not isinstance(name, str):
        return name
    name = name.strip()

    # Direct lookup
    if name in BRANCH_NAME_MAP:
        return BRANCH_NAME_MAP[name]

    # Try partial matching for Paz PDF names (which can be messy)
    name_lower = name
    for key, val in BRANCH_NAME_MAP.items():
        if key in name_lower or name_lower in key:
            return val

    return name


# ──────────────────────────────────────────────
# File classification
# ──────────────────────────────────────────────
def classify_file(file_name: str, file_bytes: bytes) -> str:
    """
    Classify uploaded file into one of:
      'csv_revenue'       – מכר כולל מעמ.csv
      'xlsx_avg_trans'     – ממוצע עסקה+מס' עסקאות.xlsx
      'xlsx_portions'      – סהכ מנות מול ארוחה בפיתה.xlsx
      'xlsx_hourly'        – עסקאות לפי שעה.xlsx
      'pdf_sales'          – מכירות חנות PDF (Paz revenue)
      'pdf_portions'       – ארוחות בפיתה PDF (Paz portions)
      'unknown'
    """
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    if ext == "csv":
        return "csv_revenue"

    if ext in ("xlsx", "xls"):
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=5)
            text = " ".join(str(v) for v in df.values.flatten() if pd.notna(v))
            if "ממוצע להזמנה" in text or "ממוצע" in text and "הזמנות" in text:
                return "xlsx_avg_trans"
            if "תמחור" in text or "ארוחה בפיתה" in text:
                return "xlsx_portions"
            if "סה\"כ" in text and ("מכירות כולל" in text or "הזמנות" in text):
                # Could be hourly transactions
                return "xlsx_hourly"
            # Fallback: check more rows
            df_full = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=20)
            text_full = " ".join(str(v) for v in df_full.values.flatten() if pd.notna(v))
            if "ממוצע" in text_full:
                return "xlsx_avg_trans"
            if "תמחור" in text_full or "נמכר" in text_full:
                return "xlsx_portions"
            return "xlsx_hourly"
        except Exception:
            return "unknown"

    if ext == "pdf":
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                first_page_text = pdf.pages[0].extract_text() or ""
                # The portions PDF has "ארוחות" and "FALAFEL_FOOD" in the header
                if "FALAFEL_FOOD" in first_page_text or "התיפב תוחורא" in first_page_text:
                    return "pdf_portions"
                # Sales PDFs have "תונח תוריכמ" (מכירות חנות reversed)
                if "תוריכמ" in first_page_text or "ןוידפ" in first_page_text:
                    return "pdf_sales"
                # Fallback: if it has transaction-like data
                if any(c.isdigit() for c in first_page_text[:200]):
                    return "pdf_sales"
        except Exception:
            pass
        return "unknown"

    return "unknown"


# ──────────────────────────────────────────────
# Parser: CSV Revenue
# ──────────────────────────────────────────────
@st.cache_data
def parse_csv_revenue(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Parse the מכר כולל מעמ.csv file → DataFrame with [סניף, מכר כולל מע"מ].

    Encoding strategy (cross-platform):
      1. utf-8-sig  – Mac/Linux default (UTF-8 with optional BOM)
      2. cp1255     – Windows Hebrew code-page (very common on Israeli Windows)
      3. utf-8      – plain UTF-8 without BOM
      4. iso-8859-8 – legacy Hebrew ISO encoding
    The first encoding that produces no UnicodeDecodeError wins.
    """
    df = None
    tried = []
    for enc in ["utf-8-sig", "cp1255", "utf-8", "iso-8859-8"]:
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc)
            # Sanity-check: at least one column should contain Hebrew characters
            sample = " ".join(str(c) for c in df.columns)
            if any("\u05d0" <= ch <= "\u05ea" for ch in sample):
                break  # Hebrew found – encoding is correct
            # No Hebrew in headers; try next encoding
            tried.append(enc)
            df = None
        except (UnicodeDecodeError, Exception):
            tried.append(enc)
            df = None

    if df is None:
        # Last resort: force-read with cp1255 ignoring errors
        try:
            df = pd.read_csv(
                io.BytesIO(file_bytes), encoding="cp1255", encoding_errors="replace"
            )
        except Exception:
            raise ValueError(
                f"לא ניתן לקרוא את קובץ ה-CSV: {file_name}\n"
                f"Tried encodings: {tried}"
            )

    # Find revenue column — MUST be column D ("בחירת מדד").
    # Column E header misleadingly says "כולל מע"מ" but contains different data.
    # Column D holds the actual gross revenue including VAT.
    rev_col = None

    # Strategy 1: look for "בחירת מדד" by name
    for c in df.columns:
        if "בחירת מדד" in str(c):
            rev_col = c
            break

    # Strategy 2: fall back to positional column D (index 3)
    if rev_col is None and len(df.columns) >= 4:
        rev_col = df.columns[3]

    # Strategy 3: last resort — pick the first numeric column with "כולל מע"
    if rev_col is None:
        for c in df.columns:
            if "כולל מע" in str(c):
                rev_col = c
                break

    if rev_col is None:
        raise ValueError("לא נמצאה עמודת מכר כולל מע\"מ בקובץ CSV")

    # Find site column
    site_col = None
    for c in df.columns:
        if "site" in str(c).lower() or "סניף" in str(c) or "חנות" in str(c):
            site_col = c
            break
    if site_col is None:
        site_col = df.columns[0]

    result = pd.DataFrame()
    result["סניף"] = df[site_col].apply(lambda x: normalize_branch(str(x)))
    result['מכר כולל מע"מ'] = pd.to_numeric(df[rev_col], errors="coerce")

    # Also extract transaction count if available
    trans_col = None
    for c in df.columns:
        if "הזמנות" in str(c):
            trans_col = c
            break
    if trans_col:
        result["הזמנות_csv"] = pd.to_numeric(df[trans_col], errors="coerce")

    result = result.dropna(subset=['מכר כולל מע"מ'])
    result = result[result["סניף"] != "Total"]
    return result


# ──────────────────────────────────────────────
# Parser: Average Transaction + Num Transactions
# ──────────────────────────────────────────────
@st.cache_data
def parse_xlsx_avg_trans(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Parse ממוצע עסקה+מס' עסקאות.xlsx → [סניף, ממוצע עסקאות, מס' עסקאות]."""
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    # Find the header row
    header_row = 0
    for i in range(min(5, len(df))):
        row_str = " ".join(str(v) for v in df.iloc[i] if pd.notna(v))
        if "סניף" in row_str or "ממוצע" in row_str:
            header_row = i
            break

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # Find relevant columns
    site_col = avg_col = trans_col = None
    for c in df.columns:
        cs = str(c)
        if "סניף" in cs or "site" in cs.lower():
            site_col = c
        elif "ממוצע" in cs:
            avg_col = c
        elif "הזמנות" in cs:
            trans_col = c

    if site_col is None:
        site_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    if avg_col is None:
        avg_col = df.columns[2] if len(df.columns) > 2 else None
    if trans_col is None:
        trans_col = df.columns[3] if len(df.columns) > 3 else None

    result = pd.DataFrame()
    result["סניף"] = df[site_col].apply(
        lambda x: normalize_branch(str(x)) if pd.notna(x) else None
    )
    if avg_col is not None:
        result["ממוצע עסקאות"] = pd.to_numeric(df[avg_col], errors="coerce")
    if trans_col is not None:
        result["מס' עסקאות"] = pd.to_numeric(df[trans_col], errors="coerce")

    result = result.dropna(subset=["סניף"])
    result = result[~result["סניף"].isin(["Total", "nan", "None"])]
    return result


# ──────────────────────────────────────────────
# Parser: Portions (Non-Paz)
# ──────────────────────────────────────────────
@st.cache_data
def parse_xlsx_portions(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Parse סהכ מנות מול ארוחה בפיתה.xlsx → [סניף, מנות בפיתה, ארוחות בפיתה]."""
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    # Structure: col0=branch, col1=category (Total/ארוחה בפיתה/מנה X), col4=qty
    # We need per-branch: Total qty and ארוחה בפיתה qty
    branch_data = {}
    current_branch = None

    for _, row in df.iterrows():
        branch_raw = str(row.iloc[0]) if pd.notna(row.iloc[0]) else None
        category = str(row.iloc[1]) if pd.notna(row.iloc[1]) else None
        qty = row.iloc[4] if len(row) > 4 and pd.notna(row.iloc[4]) else 0

        if branch_raw and branch_raw not in ("סניף", "nan", "Total", "None"):
            if "מסננים" in str(branch_raw):
                continue
            current_branch = normalize_branch(branch_raw)
            if current_branch not in branch_data:
                branch_data[current_branch] = {"total": 0, "meals": 0}

        if current_branch and category:
            qty_val = pd.to_numeric(qty, errors="coerce") or 0
            if category == "Total":
                branch_data[current_branch]["total"] = qty_val
            elif "ארוחה" in category:
                branch_data[current_branch]["meals"] = qty_val

    rows = []
    for branch, data in branch_data.items():
        total_portions = data["total"]  # This IS מנות בפיתה (includes meals)
        meals = data["meals"]
        rows.append({
            "סניף": branch,
            "מנות בפיתה": int(total_portions) if total_portions else 0,
            "ארוחות בפיתה": int(meals) if meals else 0,
        })

    return pd.DataFrame(rows)


# ──────────────────────────────────────────────
# Parser: Hourly Transactions → First/Last time
# ──────────────────────────────────────────────
@st.cache_data
def parse_xlsx_hourly(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Parse עסקאות לפי שעה.xlsx → [סניף, עסקה ראשונה, עסקה אחרונה]."""
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    branch_times = {}
    current_branch = None

    for _, row in df.iterrows():
        # Col 0: branch name (appears once per branch block)
        # Col 3: time (HH:MM)
        branch_raw = str(row.iloc[0]) if pd.notna(row.iloc[0]) else None
        time_val = row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else None

        if branch_raw and branch_raw not in ("nan", "סה\"כ", "None"):
            if "פלאפל" in branch_raw:
                current_branch = normalize_branch(branch_raw)
                if current_branch not in branch_times:
                    branch_times[current_branch] = []

        if current_branch and time_val is not None:
            time_str = str(time_val).strip()
            # Parse time
            try:
                if isinstance(time_val, datetime.time):
                    branch_times[current_branch].append(time_val)
                elif ":" in time_str:
                    parts = time_str.split(":")
                    h, m = int(parts[0]), int(parts[1])
                    branch_times[current_branch].append(datetime.time(h, m))
            except (ValueError, IndexError):
                pass

    rows = []
    for branch, times in branch_times.items():
        if times:
            first = min(times)
            last = max(times)
            rows.append({
                "סניף": branch,
                "עסקה ראשונה": first,
                "עסקה אחרונה": last,
            })

    return pd.DataFrame(rows)


# ──────────────────────────────────────────────
# Parser: PDF Sales (Paz branches)
# ──────────────────────────────────────────────
def _parse_time(s: str):
    """Try to parse a time string like '08:59:00' or '08:59'."""
    s = s.strip()
    parts = s.replace(":", " ").split()
    try:
        h, m = int(parts[0]), int(parts[1])
        return datetime.time(h, m)
    except (ValueError, IndexError):
        return None


def _parse_number(s: str) -> float:
    """Parse a number string, removing commas."""
    s = s.strip().replace(",", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


# Known Paz store names in PDF (Hebrew, may appear reversed in text extraction)
# IMPORTANT: Longer/more-specific patterns must come BEFORE shorter ones
# so that e.g. "ךז םחתמ" is matched before a bare "ךז" would be.
PAZ_STORE_PATTERNS = {
    # רופין / זך — all these variants map to רופין:
    #   "מתחם זך - ליאת חג'ג'" → extracted as "'ג'גח תאיל - ךז םחתמ"
    #   "פלאפל זך"              → extracted as "ךז לפאלפ"
    "ךז םחתמ": "רופין",
    "ךז לפאלפ": "רופין",
    "חג'ג": "רופין",
    "ליאת": "רופין",
    "פלאפל זך": "רופין",
    "ךז": "רופין",
    # טופז
    "זפוט": "טופז",
    "טופז": "טופז",
    "ינוב רחש": "טופז",
    "שחר בוני": "טופז",
    # שקמה
    "המקש": "שקמה",
    "השקמה": "שקמה",
    # מנחם
    "םחנמ": "מנחם",
    "מנחם": "מנחם",
    # אשכול
    "לוכשא": "אשכול",
    "אשכול": "אשכול",
    # נחלת
    "תלחנ": "נחלת",
    "נחלת": "נחלת",
    # פז שוהם
    "םהוש לפאלפ": "פז שוהם",
    "פלאפל שוהם": "פז שוהם",
}


def _identify_paz_store(line: str) -> str:
    """Try to identify which Paz store a line refers to."""
    for pattern, master_name in PAZ_STORE_PATTERNS.items():
        if pattern in line:
            return master_name
    return None


@st.cache_data
def parse_pdf_sales(file_bytes: bytes, file_name: str, target_date: str) -> pd.DataFrame:
    """
    Parse a Paz מכירות חנות PDF.
    Extracts: revenue incl VAT, avg, num transactions, first/last time
    for the target_date only.

    target_date: 'DD.MM.YYYY' format
    """
    results = []

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"
    except Exception as e:
        st.warning(f"שגיאה בקריאת PDF מכירות: {file_name} - {e}")
        return pd.DataFrame()

    # The PDF has store blocks. Each store section has daily rows.
    # We need to find rows matching the target date.

    # Format target_date for matching (e.g., "20.01.2026" or "20/01/2026")
    # The PDF uses DD.MM.YYYY format
    lines = all_text.split("\n")

    current_store = None

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Try to identify store from the line
        store = _identify_paz_store(line)
        if store:
            current_store = store

        # Check if this line contains our target date
        if target_date not in line:
            continue

        if current_store is None:
            # Try to get store from this line
            store = _identify_paz_store(line)
            if store:
                current_store = store
            else:
                continue

        # Extract numbers from this line
        # Pattern: numbers that could be revenue, transactions, avg, times
        # The line format (as extracted, left to right):
        # avg | num_trans | num_products | rev_excl | rev_incl | last_time | first_time | date | store

        # Find all time patterns
        time_pattern = r'(\d{1,2}:\d{2}(?::\d{2})?)'
        times = re.findall(time_pattern, line)

        # Find all number patterns (not part of time or date)
        # Remove times and date from line for number extraction
        clean_line = line
        for t in times:
            clean_line = clean_line.replace(t, " ")
        clean_line = clean_line.replace(target_date, " ")

        # Remove store name text (Hebrew chars)
        number_parts = re.findall(r'[\d,]+\.?\d*', clean_line)
        numbers = []
        for np_str in number_parts:
            val = _parse_number(np_str)
            if val is not None:
                numbers.append(val)

        if len(numbers) >= 4:
            # Expected order (left to right in extracted text):
            # avg, num_trans, num_products, rev_excl, rev_incl
            # But sometimes avg gets split. Let's use the known structure:
            # The largest number is likely revenue incl VAT
            # We can identify by relative sizes

            # Sort approach: rev_incl > rev_excl > num_products > num_trans > avg
            # Usually: numbers[-1] or numbers[-2] is rev_incl (largest)

            # Better approach: use positions
            # In the extracted text, reading left to right:
            # [avg] [num_trans] [num_products] [rev_excl_vat] [rev_incl_vat]
            avg_val = numbers[0]
            num_trans = int(numbers[1])
            rev_incl = numbers[-1]
            rev_excl = numbers[-2] if len(numbers) >= 5 else None

            # Validate: rev_incl should be roughly avg * num_trans
            if rev_incl < num_trans:
                # Numbers might be in different order, try to figure out
                # The biggest number is rev_incl
                sorted_nums = sorted(numbers, reverse=True)
                rev_incl = sorted_nums[0]
                if len(sorted_nums) > 1:
                    rev_excl = sorted_nums[1]
                # Find something close to rev_incl / something
                for n in numbers:
                    if 1 < n < 500 and rev_incl > 0:
                        potential_avg = rev_incl / n
                        if 20 < potential_avg < 150:
                            num_trans = int(n)
                            avg_val = potential_avg
                            break

            # Parse times
            first_time = None
            last_time = None
            if len(times) >= 2:
                t1 = _parse_time(times[-1])  # first_time is rightmost (RTL)
                t2 = _parse_time(times[-2])  # last_time is second from right
                if t1 and t2:
                    first_time = min(t1, t2)
                    last_time = max(t1, t2)
            elif len(times) == 1:
                first_time = _parse_time(times[0])
                last_time = first_time

            results.append({
                "סניף": current_store,
                'מכר כולל מע"מ': rev_incl,
                "ממוצע עסקאות": round(avg_val, 2),
                "מס' עסקאות": num_trans,
                "עסקה ראשונה": first_time,
                "עסקה אחרונה": last_time,
            })

            # Reset current_store for next line (each date-row is standalone)
            # Actually keep it since multiple dates for same store are sequential

    if not results:
        return pd.DataFrame()

    df = pd.DataFrame(results)
    if df.empty:
        return df

    # Aggregate: if multiple PDF lines map to the same master branch
    # (e.g. "פלאפל זך" + "מתחם זך" both → רופין), sum revenue &
    # transactions, take min/max times, and recalculate the average.
    agg_dict = {
        'מכר כולל מע"מ': "sum",
        "מס' עסקאות": "sum",
        "עסקה ראשונה": "min",
        "עסקה אחרונה": "max",
        "ממוצע עסקאות": "first",  # placeholder – recalculated below
    }
    df = df.groupby("סניף", as_index=False).agg(agg_dict)
    # Recalculate average from the summed totals
    mask = df["מס' עסקאות"] > 0
    df.loc[mask, "ממוצע עסקאות"] = (
        df.loc[mask, 'מכר כולל מע"מ'] / df.loc[mask, "מס' עסקאות"]
    ).round(2)
    return df


# ──────────────────────────────────────────────
# Parser: PDF Portions (Paz branches)
# ──────────────────────────────────────────────
@st.cache_data
def parse_pdf_portions(file_bytes: bytes, file_name: str, target_date: str) -> pd.DataFrame:
    """
    Parse the Paz ארוחות בפיתה PDF (FALAFEL_FOOD).
    For each Paz branch on target_date, extract:
      - מנות בפיתה (= non-meal portions + meals = total)
      - ארוחות בפיתה (= meals)

    target_date: 'DD.MM.YYYY' format
    """
    results = {}

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"
    except Exception as e:
        st.warning(f"שגיאה בקריאת PDF מנות: {file_name} - {e}")
        return pd.DataFrame()

    lines = all_text.split("\n")

    # Find the target date in lines
    # The PDF format has date at end of certain lines, then store lines follow
    # Each line for a store on a date:
    # [total_meals] [sub1] [sub2] [non_meal_qty] [store_name] [date]
    # OR the date appears at the end of a store line

    in_target_date_block = False

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Check if line contains our target date
        has_date = target_date in line

        # Skip summary lines (האצות = totals)
        if "האצות" in line or "תללוכ" in line:
            if has_date:
                in_target_date_block = True
            continue

        if has_date:
            in_target_date_block = True

        if not has_date and not in_target_date_block:
            continue

        # If we hit a new date, we're out of our block
        date_pattern = r'\d{2}\.\d{2}\.\d{4}'
        dates_in_line = re.findall(date_pattern, line)
        if dates_in_line and target_date not in dates_in_line:
            in_target_date_block = False
            continue

        # Try to identify the store
        store = _identify_paz_store(line)
        if store is None:
            continue

        # Extract numbers from the line
        # Remove date and store text
        clean = line
        for d in dates_in_line:
            clean = clean.replace(d, " ")

        numbers = re.findall(r'\d+', clean)
        numbers = [int(n) for n in numbers if n.strip()]

        if len(numbers) >= 2:
            # In the portions PDF, line structure (left to right as extracted):
            # [total_meals] [col1] [col2] [non_meal_qty]
            # First number (leftmost) = total meals (סה"כ ארוחות)
            # Last number before store name = non-meal portions

            total_meals = numbers[0]

            # The last "big" number that could be a qty of non-meal portions
            # is typically the number right before the store name
            # It's usually the last number in the list or close to it
            if len(numbers) >= 4:
                non_meal_qty = numbers[-1]
            elif len(numbers) == 3:
                non_meal_qty = numbers[-1]
            else:
                non_meal_qty = numbers[-1]

            # Total portions (מנות בפיתה) = non_meal + meals
            total_portions = non_meal_qty + total_meals

            # Sanity check: if non_meal_qty seems too small, it might be a sub-column
            # The non_meal_qty should generally be larger than total_meals
            # If not, try the largest number as non_meal_qty
            if non_meal_qty < total_meals and len(numbers) > 2:
                # The actual non-meal qty is likely the largest number
                candidate = max(numbers)
                if candidate > total_meals:
                    non_meal_qty = candidate
                    total_portions = non_meal_qty + total_meals

            # Accumulate: multiple PDF lines may map to the same master
            # branch (e.g. "פלאפל זך" + "מתחם זך" both → רופין).
            # Sum their portions rather than overwriting.
            if store not in results:
                results[store] = {"total_portions": 0, "meals": 0}
            results[store]["total_portions"] += total_portions
            results[store]["meals"] += total_meals

    rows = []
    for store, data in results.items():
        rows.append({
            "סניף": store,
            "מנות בפיתה": data["total_portions"],
            "ארוחות בפיתה": data["meals"],
        })

    return pd.DataFrame(rows)


# ──────────────────────────────────────────────
# Merge all data sources
# ──────────────────────────────────────────────
def merge_all(
    csv_data: pd.DataFrame,
    avg_trans_data: pd.DataFrame,
    portions_data: pd.DataFrame,
    hourly_data: pd.DataFrame,
    paz_sales_list: list,
    paz_portions_list: list,
) -> pd.DataFrame:
    """Merge all parsed data into a single DataFrame matching master format."""

    # Start with non-Paz branches from CSV
    if csv_data is not None and not csv_data.empty:
        merged = csv_data[["סניף", 'מכר כולל מע"מ']].copy()
    else:
        merged = pd.DataFrame(columns=["סניף", 'מכר כולל מע"מ'])

    # Merge average transactions
    if avg_trans_data is not None and not avg_trans_data.empty:
        avg_cols = ["סניף"]
        if "ממוצע עסקאות" in avg_trans_data.columns:
            avg_cols.append("ממוצע עסקאות")
        if "מס' עסקאות" in avg_trans_data.columns:
            avg_cols.append("מס' עסקאות")
        merged = merged.merge(avg_trans_data[avg_cols], on="סניף", how="outer")

    # Merge portions
    if portions_data is not None and not portions_data.empty:
        merged = merged.merge(
            portions_data[["סניף", "מנות בפיתה", "ארוחות בפיתה"]],
            on="סניף",
            how="outer",
        )

    # Merge hourly (first/last transaction)
    if hourly_data is not None and not hourly_data.empty:
        merged = merged.merge(
            hourly_data[["סניף", "עסקה ראשונה", "עסקה אחרונה"]],
            on="סניף",
            how="outer",
        )

    # Add Paz sales data
    for paz_sales in paz_sales_list:
        if paz_sales is not None and not paz_sales.empty:
            for _, row in paz_sales.iterrows():
                branch = row["סניף"]
                if branch in PAZ_BRANCHES:
                    # Add or update in merged
                    mask = merged["סניף"] == branch
                    if mask.any():
                        for col in row.index:
                            if col != "סניף" and pd.notna(row[col]):
                                merged.loc[mask, col] = row[col]
                    else:
                        merged = pd.concat(
                            [merged, pd.DataFrame([row])], ignore_index=True
                        )

    # Add Paz portions data
    for paz_portions in paz_portions_list:
        if paz_portions is not None and not paz_portions.empty:
            for _, row in paz_portions.iterrows():
                branch = row["סניף"]
                if branch in PAZ_BRANCHES:
                    mask = merged["סניף"] == branch
                    if mask.any():
                        if pd.notna(row.get("מנות בפיתה")):
                            merged.loc[mask, "מנות בפיתה"] = row["מנות בפיתה"]
                        if pd.notna(row.get("ארוחות בפיתה")):
                            merged.loc[mask, "ארוחות בפיתה"] = row["ארוחות בפיתה"]
                    else:
                        new_row = {
                            "סניף": branch,
                            "מנות בפיתה": row.get("מנות בפיתה", 0),
                            "ארוחות בפיתה": row.get("ארוחות בפיתה", 0),
                        }
                        merged = pd.concat(
                            [merged, pd.DataFrame([new_row])], ignore_index=True
                        )

    # Calculate meal percentage
    merged["אחוז ארוחות מתוך מנות"] = merged.apply(
        lambda r: (
            r["ארוחות בפיתה"] / r["מנות בפיתה"]
            if pd.notna(r.get("מנות בפיתה"))
            and pd.notna(r.get("ארוחות בפיתה"))
            and r["מנות בפיתה"] > 0
            else 0
        ),
        axis=1,
    )

    # Ensure all master columns exist
    for col in MASTER_COLUMNS:
        if col not in merged.columns:
            merged[col] = None

    # Reorder columns
    merged = merged[MASTER_COLUMNS]

    # Sort by master branch order
    order_map = {b: i for i, b in enumerate(MASTER_BRANCH_ORDER)}
    merged["_sort"] = merged["סניף"].map(order_map).fillna(99)
    merged = merged.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

    # Filter out non-branch rows
    merged = merged[merged["סניף"].notna() & (merged["סניף"] != "")]

    return merged


# ──────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────
def generate_excel(df: pd.DataFrame, report_date: datetime.date) -> bytes:
    """Generate a professionally formatted Excel report.

    Layout (compact, no wasted rows):
      Row 1 : Title bar — date + report name, merged across all columns
      Row 2 : Column headers (blue background, white bold text)
      Row 3…: Data rows (alternating zebra stripes)
      Last  : Totals row (bold, light-blue background)

    Uses only in-memory io.BytesIO — works identically on Windows and macOS.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    # Sheet name: DD.M format (like "15.2")
    ws.title = f"{report_date.day}.{report_date.month}"

    # ── Styles ──
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(size=11)
    branch_font = Font(size=11, bold=True)
    totals_font = Font(bold=True, size=11, color="1F3864")
    totals_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    stripe_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    num_cols = len(MASTER_COLUMNS)

    # ── Row 1: Title bar (merged) ──
    date_str = report_date.strftime("%d.%m.%Y")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"דוח מכירות יומי — {date_str}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 32

    # Fill the rest of the merged row so the border/fill extends visually
    for ci in range(2, num_cols + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = title_fill
        c.border = thin_border

    # ── Row 2: Column headers ──
    ws.row_dimensions[2].height = 28
    for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # ── Prepare export data ──
    export_df = df.copy()

    # Format time columns as HH:MM strings
    for col in ["עסקה ראשונה", "עסקה אחרונה"]:
        export_df[col] = export_df[col].apply(
            lambda x: x.strftime("%H:%M") if isinstance(x, datetime.time) else (
                str(x) if pd.notna(x) else ""
            )
        )

    # ── Rows 3…N: Data ──
    data_start_row = 3
    for ri, (_, row) in enumerate(export_df.iterrows()):
        excel_row = data_start_row + ri
        is_stripe = ri % 2 == 1  # alternate rows

        for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
            val = row.get(col_name)
            cell = ws.cell(row=excel_row, column=ci)
            cell.border = thin_border

            if pd.isna(val):
                cell.value = None
            else:
                cell.value = val

            # Styling per column
            if col_name == "סניף":
                cell.font = branch_font
                cell.alignment = right_align
            elif col_name == 'מכר כולל מע"מ':
                cell.number_format = "#,##0"
                cell.font = data_font
                cell.alignment = center
            elif col_name == "ממוצע עסקאות":
                cell.number_format = "#,##0.00"
                cell.font = data_font
                cell.alignment = center
            elif col_name == "אחוז ארוחות מתוך מנות":
                cell.number_format = "0.00%"
                cell.font = data_font
                cell.alignment = center
            elif col_name in ("מנות בפיתה", "ארוחות בפיתה", "מס' עסקאות"):
                cell.number_format = "#,##0"
                cell.font = data_font
                cell.alignment = center
            else:
                cell.font = data_font
                cell.alignment = center

            # Zebra stripe
            if is_stripe:
                cell.fill = stripe_fill

    # ── Totals row ──
    totals_row_idx = data_start_row + len(export_df)
    totals_values = {
        "סניף": 'סה"כ',
        'מכר כולל מע"מ': df['מכר כולל מע"מ'].sum(),
        "ממוצע עסקאות": None,
        "מס' עסקאות": df["מס' עסקאות"].sum() if "מס' עסקאות" in df.columns else None,
        "עסקה ראשונה": None,
        "עסקה אחרונה": None,
        "מנות בפיתה": df["מנות בפיתה"].sum(),
        "ארוחות בפיתה": df["ארוחות בפיתה"].sum(),
    }
    total_p = totals_values["מנות בפיתה"] or 0
    total_m = totals_values["ארוחות בפיתה"] or 0
    totals_values["אחוז ארוחות מתוך מנות"] = total_m / total_p if total_p > 0 else 0

    for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=totals_row_idx, column=ci, value=totals_values.get(col_name))
        cell.font = totals_font
        cell.fill = totals_fill
        cell.alignment = center
        cell.border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="medium", color="4472C4"),
            bottom=Side(style="medium", color="4472C4"),
        )
        # Number formatting for totals
        if col_name == 'מכר כולל מע"מ':
            cell.number_format = "#,##0"
        elif col_name == "אחוז ארוחות מתוך מנות":
            cell.number_format = "0.00%"
        elif col_name in ("מנות בפיתה", "ארוחות בפיתה", "מס' עסקאות"):
            cell.number_format = "#,##0"

    # ── Column widths (generous for Hebrew text) ──
    col_widths = {
        "סניף": 16,
        'מכר כולל מע"מ': 16,
        "ממוצע עסקאות": 15,
        "מס' עסקאות": 13,
        "עסקה ראשונה": 15,
        "עסקה אחרונה": 15,
        "מנות בפיתה": 14,
        "ארוחות בפיתה": 15,
        "אחוז ארוחות מתוך מנות": 22,
    }
    for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 14)

    # ── Freeze panes: keep title + headers visible when scrolling ──
    ws.freeze_panes = "A3"

    # ── RTL sheet direction ──
    ws.sheet_view.rightToLeft = True

    # ── Print settings (nice printout) ──
    ws.print_title_rows = "1:2"  # repeat title+headers on every printed page
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ══════════════════════════════════════════════
    # Sheet 2: "ניתוח" (Analysis)
    # ══════════════════════════════════════════════
    _build_analysis_sheet(wb, df, report_date, ws.title)

    wb.save(output)
    return output.getvalue()


def _build_analysis_sheet(wb, df: pd.DataFrame, report_date: datetime.date, data_sheet_name: str):
    """Build a second sheet with derived analytics, rankings, and a chart.

    Contents:
      - Section A: Branch ranking table (sorted by revenue) with extra
        derived columns: revenue %, operating hours, revenue/hour
      - Section B: Paz vs Non-Paz group summary
      - Section C: Embedded bar chart — revenue by branch
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    ws2 = wb.create_sheet(title="ניתוח")
    ws2.sheet_view.rightToLeft = True

    # ── Shared styles ──
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(size=10)
    bold_font = Font(size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right_a = Alignment(horizontal="right", vertical="center")
    stripe = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    subtotal_font = Font(bold=True, size=10, color="1F3864")

    num_branches = len(df)
    total_rev = df['מכר כולל מע"מ'].sum()

    # ── Prepare analysis DataFrame (sorted by revenue descending) ──
    adf = df.copy().sort_values('מכר כולל מע"מ', ascending=False).reset_index(drop=True)

    # Revenue %
    adf["% מהמכר"] = adf['מכר כולל מע"מ'] / total_rev if total_rev > 0 else 0

    # Operating hours (from first to last transaction)
    def _calc_hours(row):
        f = row.get("עסקה ראשונה")
        l = row.get("עסקה אחרונה")
        if isinstance(f, datetime.time) and isinstance(l, datetime.time):
            f_min = f.hour * 60 + f.minute
            l_min = l.hour * 60 + l.minute
            diff = l_min - f_min
            return round(diff / 60, 1) if diff > 0 else None
        return None
    adf["שעות פעילות"] = adf.apply(_calc_hours, axis=1)

    # Revenue per hour
    adf["מכר לשעה"] = adf.apply(
        lambda r: round(r['מכר כולל מע"מ'] / r["שעות פעילות"])
        if r["שעות פעילות"] and r["שעות פעילות"] > 0 else None,
        axis=1,
    )

    # Is Paz?
    adf["_paz"] = adf["סניף"].isin(PAZ_BRANCHES)

    # ────────────────────────────────
    # Row 1: Title
    # ────────────────────────────────
    analysis_cols = [
        "דירוג", "סניף", 'מכר כולל מע"מ', "% מהמכר",
        "מס' עסקאות", "שעות פעילות", "מכר לשעה",
        "מנות בפיתה", "ארוחות בפיתה",
    ]
    ncols = len(analysis_cols)

    date_str = report_date.strftime("%d.%m.%Y")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws2.cell(row=1, column=1, value=f"ניתוח יומי — {date_str}")
    tc.font = title_font
    tc.fill = title_fill
    tc.alignment = center
    tc.border = thin_border
    ws2.row_dimensions[1].height = 32
    for ci in range(2, ncols + 1):
        c = ws2.cell(row=1, column=ci)
        c.fill = title_fill
        c.border = thin_border

    # ────────────────────────────────
    # Section A: Ranked Branch Table
    # ────────────────────────────────
    sec_a_row = 3
    ws2.cell(row=sec_a_row, column=1, value="דירוג סניפים לפי מכר").font = section_font

    hdr_row = sec_a_row + 1
    ws2.row_dimensions[hdr_row].height = 26
    for ci, col_name in enumerate(analysis_cols, start=1):
        cell = ws2.cell(row=hdr_row, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    data_start = hdr_row + 1
    for ri, (_, row) in enumerate(adf.iterrows()):
        er = data_start + ri
        rank = ri + 1
        is_stripe = ri % 2 == 1

        values = [
            rank,
            row["סניף"],
            row['מכר כולל מע"מ'],
            row["% מהמכר"],
            row.get("מס' עסקאות"),
            row.get("שעות פעילות"),
            row.get("מכר לשעה"),
            row.get("מנות בפיתה"),
            row.get("ארוחות בפיתה"),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws2.cell(row=er, column=ci)
            cell.value = val if pd.notna(val) else None
            cell.border = thin_border
            cell.alignment = center
            cell.font = bold_font if ci == 2 else data_font

            # Number formats
            col_name = analysis_cols[ci - 1]
            if col_name == 'מכר כולל מע"מ':
                cell.number_format = "#,##0"
            elif col_name == "% מהמכר":
                cell.number_format = "0.0%"
            elif col_name == "מכר לשעה":
                cell.number_format = "#,##0"
            elif col_name in ("מנות בפיתה", "ארוחות בפיתה", "מס' עסקאות"):
                cell.number_format = "#,##0"
            elif col_name == "שעות פעילות":
                cell.number_format = "0.0"

            if is_stripe:
                cell.fill = stripe

        # Highlight top 3 with gold, bottom 3 with red
        if rank <= 3:
            for ci in range(1, ncols + 1):
                ws2.cell(row=er, column=ci).fill = gold_fill
        elif rank > num_branches - 3:
            for ci in range(1, ncols + 1):
                ws2.cell(row=er, column=ci).fill = red_fill

    # ────────────────────────────────
    # Section B: Paz vs Non-Paz Summary
    # ────────────────────────────────
    sec_b_row = data_start + num_branches + 2
    ws2.cell(row=sec_b_row, column=1, value="סיכום: רשת מול פז").font = section_font

    summary_cols = ["קבוצה", "סניפים", 'מכר כולל מע"מ', "% מהמכר", "מס' עסקאות", "ממוצע מכר לסניף"]
    shdr = sec_b_row + 1
    ws2.row_dimensions[shdr].height = 26
    for ci, cn in enumerate(summary_cols, start=1):
        cell = ws2.cell(row=shdr, column=ci, value=cn)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    paz_df = adf[adf["_paz"]]
    non_paz_df = adf[~adf["_paz"]]

    groups = [
        ("סניפי רשת", non_paz_df),
        ("סניפי פז", paz_df),
        ('סה"כ', adf),
    ]
    for gi, (label, gdf) in enumerate(groups):
        r = shdr + 1 + gi
        g_rev = gdf['מכר כולל מע"מ'].sum()
        g_count = len(gdf)
        g_trans = gdf["מס' עסקאות"].sum() if "מס' עסקאות" in gdf.columns else 0
        g_pct = g_rev / total_rev if total_rev > 0 else 0
        g_avg = g_rev / g_count if g_count > 0 else 0

        vals = [label, g_count, g_rev, g_pct, g_trans, g_avg]
        is_total = (gi == 2)
        for ci, val in enumerate(vals, start=1):
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = center
            cell.font = subtotal_font if is_total else bold_font
            if is_total:
                cell.fill = subtotal_fill
            elif gi == 0:
                cell.fill = green_fill

            cn = summary_cols[ci - 1]
            if cn == 'מכר כולל מע"מ':
                cell.number_format = "#,##0"
            elif cn == "% מהמכר":
                cell.number_format = "0.0%"
            elif cn == "ממוצע מכר לסניף":
                cell.number_format = "#,##0"
            elif cn == "מס' עסקאות":
                cell.number_format = "#,##0"

    # ────────────────────────────────
    # Section C: Charts
    # ────────────────────────────────
    from openpyxl.chart import PieChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList

    # Charts are laid out in a single vertical column with generous
    # spacing so they never overlap — critical for Google Sheets compat.
    # At default row height (~0.5 cm / ~15 px), a 14 cm chart needs
    # ~28 rows.  We use CHART_GAP = 34 rows to be safe.
    CHART_H   = 14          # cm – height for full-width bar charts
    CHART_W   = 24          # cm – width for full-width bar charts
    SMALL_H   = 12          # cm – height for smaller charts (pie)
    SMALL_W   = 18          # cm – width for smaller charts (pie)
    CHART_GAP = 34          # rows between chart anchors

    from openpyxl.chart import PieChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList

    # Helper data area starts far to the right (col 14+) so it's out
    # of the visible table columns but still available as chart source.
    HELPER_COL = 14

    chart_row = shdr + 5  # first chart anchor row

    # ── Chart 1: Revenue by branch (horizontal bar) ──
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = "מכר לפי סניף"
    chart1.y_axis.title = None
    chart1.x_axis.title = None
    chart1.style = 26
    chart1.width = CHART_W
    chart1.height = CHART_H
    chart1.legend = None

    data_ref = Reference(ws2, min_col=3, min_row=hdr_row,
                         max_row=data_start + num_branches - 1, max_col=3)
    cats_ref = Reference(ws2, min_col=2, min_row=data_start,
                         max_row=data_start + num_branches - 1)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.solidFill = "4472C4"
    ws2.add_chart(chart1, f"A{chart_row}")

    # ── Chart 2: Paz vs Non-Paz pie chart ──
    chart_row += CHART_GAP

    # Helper data for pie
    pie_data_row = chart_row
    ws2.cell(row=pie_data_row, column=HELPER_COL, value="קבוצה")
    ws2.cell(row=pie_data_row, column=HELPER_COL + 1, value="מכר")
    paz_rev = adf.loc[adf["_paz"], 'מכר כולל מע"מ'].sum()
    non_paz_rev = adf.loc[~adf["_paz"], 'מכר כולל מע"מ'].sum()
    ws2.cell(row=pie_data_row + 1, column=HELPER_COL, value="סניפי רשת")
    ws2.cell(row=pie_data_row + 1, column=HELPER_COL + 1, value=non_paz_rev)
    ws2.cell(row=pie_data_row + 2, column=HELPER_COL, value="סניפי פז")
    ws2.cell(row=pie_data_row + 2, column=HELPER_COL + 1, value=paz_rev)

    pie = PieChart()
    pie.title = "פילוח מכר: רשת מול פז"
    pie.style = 26
    pie.width = SMALL_W
    pie.height = SMALL_H

    pie_data = Reference(ws2, min_col=HELPER_COL + 1, min_row=pie_data_row,
                         max_row=pie_data_row + 2)
    pie_cats = Reference(ws2, min_col=HELPER_COL, min_row=pie_data_row + 1,
                         max_row=pie_data_row + 2)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)

    slice_chain = DataPoint(idx=0)
    slice_chain.graphicalProperties.solidFill = "4472C4"
    slice_paz = DataPoint(idx=1)
    slice_paz.graphicalProperties.solidFill = "F2994A"
    pie.series[0].data_points = [slice_chain, slice_paz]

    pie.series[0].dLbls = DataLabelList()
    pie.series[0].dLbls.showPercent = True
    pie.series[0].dLbls.showCatName = True
    pie.series[0].dLbls.showVal = False

    ws2.add_chart(pie, f"A{chart_row}")

    # ── Chart 3: Revenue per hour (efficiency) ──
    chart_row += CHART_GAP

    rph_header_row = chart_row
    ws2.cell(row=rph_header_row, column=HELPER_COL, value="סניף")
    ws2.cell(row=rph_header_row, column=HELPER_COL + 1, value="מכר לשעה")
    rph_df = adf[adf["מכר לשעה"].notna()].sort_values("מכר לשעה", ascending=False)
    for ri, (_, row) in enumerate(rph_df.iterrows()):
        ws2.cell(row=rph_header_row + 1 + ri, column=HELPER_COL, value=row["סניף"])
        ws2.cell(row=rph_header_row + 1 + ri, column=HELPER_COL + 1, value=row["מכר לשעה"])
    rph_count = len(rph_df)

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "מכר לשעת פעילות"
    chart3.y_axis.title = None
    chart3.x_axis.title = None
    chart3.style = 26
    chart3.width = CHART_W
    chart3.height = CHART_H
    chart3.legend = None

    rph_data_ref = Reference(ws2, min_col=HELPER_COL + 1, min_row=rph_header_row,
                             max_row=rph_header_row + rph_count)
    rph_cats_ref = Reference(ws2, min_col=HELPER_COL, min_row=rph_header_row + 1,
                             max_row=rph_header_row + rph_count)
    chart3.add_data(rph_data_ref, titles_from_data=True)
    chart3.set_categories(rph_cats_ref)
    chart3.series[0].graphicalProperties.solidFill = "11998E"
    ws2.add_chart(chart3, f"A{chart_row}")

    # ── Chart 4: Meal % of portions ──
    chart_row += CHART_GAP

    meal_header_row = chart_row
    ws2.cell(row=meal_header_row, column=HELPER_COL, value="סניף")
    ws2.cell(row=meal_header_row, column=HELPER_COL + 1, value="% ארוחות")
    meal_df = adf[
        (adf["מנות בפיתה"].notna()) & (adf["מנות בפיתה"] > 0) &
        (adf["ארוחות בפיתה"].notna())
    ].copy()
    meal_df["_meal_pct"] = meal_df["ארוחות בפיתה"] / meal_df["מנות בפיתה"]
    meal_df = meal_df.sort_values("_meal_pct", ascending=False)
    for ri, (_, row) in enumerate(meal_df.iterrows()):
        ws2.cell(row=meal_header_row + 1 + ri, column=HELPER_COL, value=row["סניף"])
        ws2.cell(row=meal_header_row + 1 + ri, column=HELPER_COL + 1, value=row["_meal_pct"])
        ws2.cell(row=meal_header_row + 1 + ri, column=HELPER_COL + 1).number_format = "0.0%"
    meal_count = len(meal_df)

    chart4 = BarChart()
    chart4.type = "bar"
    chart4.title = "אחוז ארוחות מתוך מנות בפיתה"
    chart4.y_axis.title = None
    chart4.x_axis.title = None
    chart4.style = 26
    chart4.width = CHART_W
    chart4.height = CHART_H
    chart4.legend = None

    meal_data_ref = Reference(ws2, min_col=HELPER_COL + 1, min_row=meal_header_row,
                              max_row=meal_header_row + meal_count)
    meal_cats_ref = Reference(ws2, min_col=HELPER_COL, min_row=meal_header_row + 1,
                              max_row=meal_header_row + meal_count)
    chart4.add_data(meal_data_ref, titles_from_data=True)
    chart4.set_categories(meal_cats_ref)
    chart4.series[0].graphicalProperties.solidFill = "F2C94C"
    ws2.add_chart(chart4, f"A{chart_row}")

    # ── Column widths ──
    widths = [8, 16, 16, 12, 13, 13, 14, 14, 14]
    for ci, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = "A5"


# ──────────────────────────────────────────────
# Main Streamlit App
# ──────────────────────────────────────────────
def main():
    # Custom CSS
    st.markdown(
        """
        <style>
        .main-header {
            text-align: center;
            padding: 0.5rem 0;
        }
        .metric-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 1.2rem;
            border-radius: 12px;
            color: white;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
        .metric-card h3 {
            font-size: 0.9rem;
            margin-bottom: 0.3rem;
            opacity: 0.9;
        }
        .metric-card h1 {
            font-size: 2rem;
            margin: 0;
            font-weight: 700;
        }
        .metric-green {
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        }
        .metric-orange {
            background: linear-gradient(135deg, #F2994A 0%, #F2C94C 100%);
        }
        .metric-blue {
            background: linear-gradient(135deg, #2193b0 0%, #6dd5ed 100%);
        }
        .metric-card-big {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 1.5rem;
            border-radius: 12px;
            color: white;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
        .metric-card-big h3 {
            font-size: 1.1rem;
            margin-bottom: 0.3rem;
            opacity: 0.9;
        }
        .metric-card-big h1 {
            font-size: 2.8rem;
            margin: 0;
            font-weight: 700;
        }
        .section-header {
            font-size: 1.4rem;
            font-weight: 700;
            color: #2F5496;
            margin-top: 1.5rem;
            margin-bottom: 0.5rem;
            border-bottom: 2px solid #4472C4;
            padding-bottom: 0.4rem;
        }
        .group-card {
            padding: 1rem;
            border-radius: 10px;
            text-align: center;
            border: 2px solid #ddd;
        }
        .group-card h4 { font-size: 0.85rem; margin-bottom: 0.2rem; color: #555; }
        .group-card h2 { font-size: 1.6rem; margin: 0.2rem 0; font-weight: 700; }
        .group-card p  { font-size: 1.1rem; margin: 0; color: #666; }
        .group-chain { border-color: #4472C4; background: #f0f5ff; }
        .group-paz   { border-color: #F2994A; background: #fff8f0; }
        .file-ok   { color: #28a745; }
        .file-miss { color: #dc3545; }
        .stDownloadButton > button {
            width: 100%;
            background-color: #4CAF50 !important;
            color: white !important;
            font-size: 1.3rem !important;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            border: none !important;
            font-weight: bold !important;
        }
        .stDownloadButton > button:hover {
            background-color: #45a049 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ── Sidebar ──
    with st.sidebar:
        st.image(
            "https://img.icons8.com/color/96/falafel-in-pita.png",
            width=80,
        )
        st.title("פלאפל בריבוע")
        st.caption("דוח מכירות יומי")
        st.divider()

        # Platform info (useful for debugging encoding on Windows vs Mac)
        _platform = "🪟 Windows" if os.name == "nt" else "🍎 macOS / Linux"
        st.caption(f"מערכת הפעלה: {_platform}")

        report_date = st.date_input(
            "📅 תאריך הדוח",
            value=datetime.date.today(),
            format="DD/MM/YYYY",
        )

        st.divider()

        uploaded_files = st.file_uploader(
            "📁 העלאת קבצים",
            type=["csv", "xlsx", "xls", "pdf"],
            accept_multiple_files=True,
            help="העלה את כל קבצי היום: CSV, Excel ו-PDF",
        )

        if uploaded_files:
            st.success(f"✅ {len(uploaded_files)} קבצים הועלו")

        # Placeholder for download button — filled after data is processed
        sidebar_download_placeholder = st.empty()

    # ── Main area ──
    st.markdown(
        "<h1 class='main-header'>📊 דוח מכירות יומי - פלאפל בריבוע</h1>",
        unsafe_allow_html=True,
    )

    if not uploaded_files:
        st.info(
            "👈 העלה קבצים בסרגל הצד כדי להתחיל.\n\n"
            "**קבצים נדרשים:**\n"
            "- 🟢 CSV מכר כולל מע\"מ\n"
            "- 🟢 Excel ממוצע עסקה + מס' עסקאות\n"
            "- 🟢 Excel סה\"כ מנות מול ארוחה בפיתה\n"
            "- 🟢 Excel עסקאות לפי שעה\n"
            "- 🟠 PDF מכירות חנות (סניפי פז)\n"
            "- 🟠 PDF ארוחות בפיתה (סניפי פז)"
        )
        return

    # ── Process files ──
    target_date_str = report_date.strftime("%d.%m.%Y")

    # Classify files
    csv_data = None
    avg_trans_data = None
    portions_data = None
    hourly_data = None
    paz_sales_list = []
    paz_portions_list = []
    file_classifications = {}

    for f in uploaded_files:
        file_bytes = f.read()
        f.seek(0)  # Reset for potential re-read
        ftype = classify_file(f.name, file_bytes)
        file_classifications[f.name] = ftype

        try:
            if ftype == "csv_revenue":
                csv_data = parse_csv_revenue(file_bytes, f.name)
            elif ftype == "xlsx_avg_trans":
                avg_trans_data = parse_xlsx_avg_trans(file_bytes, f.name)
            elif ftype == "xlsx_portions":
                portions_data = parse_xlsx_portions(file_bytes, f.name)
            elif ftype == "xlsx_hourly":
                hourly_data = parse_xlsx_hourly(file_bytes, f.name)
            elif ftype == "pdf_sales":
                result = parse_pdf_sales(file_bytes, f.name, target_date_str)
                if result is not None and not result.empty:
                    paz_sales_list.append(result)
            elif ftype == "pdf_portions":
                result = parse_pdf_portions(file_bytes, f.name, target_date_str)
                if result is not None and not result.empty:
                    paz_portions_list.append(result)
            else:
                st.warning(f"⚠️ לא הצלחתי לזהות את סוג הקובץ: {f.name}")
        except Exception as e:
            st.error(f"❌ שגיאה בעיבוד {f.name}: {e}")

    # Show file completeness checklist
    _detected_types = set(file_classifications.values())
    _required_files = [
        ("csv_revenue", "CSV מכר כולל מע\"מ"),
        ("xlsx_avg_trans", "Excel ממוצע עסקה"),
        ("xlsx_portions", "Excel מנות בפיתה"),
        ("xlsx_hourly", "Excel עסקאות לפי שעה"),
    ]
    _optional_files = [
        ("pdf_sales", "PDF מכירות פז"),
        ("pdf_portions", "PDF מנות פז"),
    ]
    _missing_required = [label for ftype, label in _required_files if ftype not in _detected_types]
    if _missing_required:
        st.warning("⚠️ **קבצים חסרים:** " + " · ".join(_missing_required))

    with st.expander("🔍 זיהוי קבצים", expanded=bool(_missing_required)):
        type_labels = {
            "csv_revenue": "CSV מכר כולל מע\"מ",
            "xlsx_avg_trans": "Excel ממוצע עסקה",
            "xlsx_portions": "Excel מנות בפיתה",
            "xlsx_hourly": "Excel עסקאות לפי שעה",
            "pdf_sales": "PDF מכירות חנות (פז)",
            "pdf_portions": "PDF ארוחות בפיתה (פז)",
            "unknown": "❓ לא ידוע",
        }
        for ftype, label in _required_files + _optional_files:
            icon = "✅" if ftype in _detected_types else "❌"
            st.markdown(f"{icon} {label}")
        st.divider()
        for fname, ftype in file_classifications.items():
            st.write(f"**{fname}** → {type_labels.get(ftype, ftype)}")

    # Merge everything
    try:
        merged = merge_all(
            csv_data, avg_trans_data, portions_data, hourly_data,
            paz_sales_list, paz_portions_list,
        )
    except Exception as e:
        st.error(f"❌ שגיאה במיזוג הנתונים: {e}")
        return

    if merged.empty:
        st.warning("לא נמצאו נתונים לתאריך הנבחר. ודא שהקבצים תואמים את התאריך.")
        return

    # ── KPI Cards ──
    st.markdown("---")

    total_revenue = merged['מכר כולל מע"מ'].sum()
    total_portions = merged["מנות בפיתה"].sum()
    total_transactions = merged["מס' עסקאות"].sum()
    top_branch = merged.loc[merged['מכר כולל מע"מ'].idxmax(), "סניף"] if not merged['מכר כולל מע"מ'].isna().all() else "N/A"
    top_revenue = merged['מכר כולל מע"מ'].max()
    num_active = len(merged[merged['מכר כולל מע"מ'] > 0])

    # Row 1: Revenue as the visually dominant KPI
    st.markdown(
        f"""
        <div class="metric-card-big">
            <h3>מכר כולל יומי</h3>
            <h1>₪{total_revenue:,.0f}</h1>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # Row 2: Secondary KPIs
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(
            f"""
            <div class="metric-card metric-green">
                <h3>סה"כ מנות</h3>
                <h1>{total_portions:,.0f}</h1>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col2:
        st.markdown(
            f"""
            <div class="metric-card metric-orange">
                <h3>סה"כ עסקאות</h3>
                <h1>{total_transactions:,.0f}</h1>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col3:
        st.markdown(
            f"""
            <div class="metric-card metric-blue">
                <h3>סניף מוביל</h3>
                <h1 style="font-size:1.4rem">{top_branch}</h1>
                <p style="margin:0;opacity:0.8">₪{top_revenue:,.0f}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col4:
        st.markdown(
            f"""
            <div class="metric-card">
                <h3>סניפים פעילים</h3>
                <h1>{num_active}</h1>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════
    # 📈 Analytics Section — mirrors Excel "ניתוח" sheet
    # ══════════════════════════════════════════════
    st.markdown('<div class="section-header">📈 ניתוח יומי</div>', unsafe_allow_html=True)

    # ── Paz vs Non-Paz split (metric cards) ──
    _paz_mask = merged["סניף"].isin(PAZ_BRANCHES)
    paz_rev = merged.loc[_paz_mask, 'מכר כולל מע"מ'].sum()
    non_paz_rev = merged.loc[~_paz_mask, 'מכר כולל מע"מ'].sum()
    paz_pct = paz_rev / total_revenue * 100 if total_revenue > 0 else 0
    non_paz_pct = non_paz_rev / total_revenue * 100 if total_revenue > 0 else 0
    paz_count = _paz_mask.sum()
    non_paz_count = (~_paz_mask).sum()

    gcol1, gcol2 = st.columns(2)
    with gcol1:
        st.markdown(
            f"""
            <div class="group-card group-chain">
                <h4>סניפי רשת ({non_paz_count})</h4>
                <h2>₪{non_paz_rev:,.0f}</h2>
                <p>{non_paz_pct:.1f}% מהמכר</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with gcol2:
        st.markdown(
            f"""
            <div class="group-card group-paz">
                <h4>סניפי פז ({paz_count})</h4>
                <h2>₪{paz_rev:,.0f}</h2>
                <p>{paz_pct:.1f}% מהמכר</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Branch ranking table ──
    st.markdown("**דירוג סניפים לפי מכר**")

    ranking_df = merged.copy()
    ranking_df = ranking_df.sort_values('מכר כולל מע"מ', ascending=False).reset_index(drop=True)

    # Compute derived columns
    ranking_df["דירוג"] = range(1, len(ranking_df) + 1)
    ranking_df["% מהמכר"] = ranking_df['מכר כולל מע"מ'] / total_revenue if total_revenue > 0 else 0

    def _calc_hours(row):
        f = row.get("עסקה ראשונה")
        l = row.get("עסקה אחרונה")
        if isinstance(f, datetime.time) and isinstance(l, datetime.time):
            f_min = f.hour * 60 + f.minute
            l_min = l.hour * 60 + l.minute
            diff = l_min - f_min
            return round(diff / 60, 1) if diff > 0 else None
        return None

    ranking_df["שעות פעילות"] = ranking_df.apply(_calc_hours, axis=1)
    ranking_df["מכר לשעה"] = ranking_df.apply(
        lambda r: round(r['מכר כולל מע"מ'] / r["שעות פעילות"])
        if r["שעות פעילות"] and r["שעות פעילות"] > 0 else None,
        axis=1,
    )

    # Build display version
    rank_display = ranking_df[["דירוג", "סניף", 'מכר כולל מע"מ', "% מהמכר",
                                "מס' עסקאות", "שעות פעילות", "מכר לשעה",
                                "מנות בפיתה", "ארוחות בפיתה"]].copy()

    n_branches = len(rank_display)

    def _highlight_rank(row):
        rank = row["דירוג"]
        if rank <= 3:
            return ["background-color: #FFF2CC"] * len(row)
        elif rank > n_branches - 3:
            return ["background-color: #FCE4EC"] * len(row)
        return [""] * len(row)

    styled_ranking = rank_display.style.apply(_highlight_rank, axis=1).format({
        'מכר כולל מע"מ': "₪{:,.0f}",
        "% מהמכר": "{:.1%}",
        "מס' עסקאות": lambda x: f"{int(x)}" if pd.notna(x) else "",
        "שעות פעילות": lambda x: f"{x:.1f}" if pd.notna(x) else "",
        "מכר לשעה": lambda x: f"₪{x:,.0f}" if pd.notna(x) else "",
        "מנות בפיתה": lambda x: f"{int(x)}" if pd.notna(x) else "",
        "ארוחות בפיתה": lambda x: f"{int(x)}" if pd.notna(x) else "",
    })

    st.dataframe(
        styled_ranking,
        use_container_width=True,
        hide_index=True,
        height=min(40 * n_branches + 60, 600),
    )

    # ── Paz vs Non-Paz summary table ──
    st.markdown("**סיכום: רשת מול פז**")

    paz_trans = merged.loc[_paz_mask, "מס' עסקאות"].sum()
    non_paz_trans = merged.loc[~_paz_mask, "מס' עסקאות"].sum()

    summary_rows = [
        {
            "קבוצה": "סניפי רשת",
            "סניפים": int(non_paz_count),
            'מכר כולל מע"מ': non_paz_rev,
            "% מהמכר": non_paz_pct / 100,
            "מס' עסקאות": int(non_paz_trans),
            "ממוצע מכר לסניף": round(non_paz_rev / non_paz_count) if non_paz_count > 0 else 0,
        },
        {
            "קבוצה": "סניפי פז",
            "סניפים": int(paz_count),
            'מכר כולל מע"מ': paz_rev,
            "% מהמכר": paz_pct / 100,
            "מס' עסקאות": int(paz_trans),
            "ממוצע מכר לסניף": round(paz_rev / paz_count) if paz_count > 0 else 0,
        },
        {
            "קבוצה": 'סה"כ',
            "סניפים": int(non_paz_count + paz_count),
            'מכר כולל מע"מ': total_revenue,
            "% מהמכר": 1.0,
            "מס' עסקאות": int(non_paz_trans + paz_trans),
            "ממוצע מכר לסניף": round(total_revenue / (non_paz_count + paz_count)) if (non_paz_count + paz_count) > 0 else 0,
        },
    ]
    summary_df = pd.DataFrame(summary_rows)

    st.dataframe(
        summary_df.style.format({
            'מכר כולל מע"מ': "₪{:,.0f}",
            "% מהמכר": "{:.1%}",
            "מס' עסקאות": "{:,.0f}",
            "ממוצע מכר לסניף": "₪{:,.0f}",
        }),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts (expanded from 2 to 4 tabs) ──
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 מכר לפי סניף",
        "⏱️ מכר לשעת פעילות",
        "🥙 מנות וארוחות",
        "📈 אחוז ארוחות",
    ])

    with tab1:
        chart_df = merged[["סניף", 'מכר כולל מע"מ']].copy()
        chart_df = chart_df.sort_values('מכר כולל מע"מ', ascending=True)
        st.bar_chart(
            chart_df.set_index("סניף")['מכר כולל מע"מ'],
            horizontal=True,
            use_container_width=True,
        )

    with tab2:
        rph_chart = ranking_df[["סניף", "מכר לשעה"]].copy()
        rph_chart = rph_chart.dropna(subset=["מכר לשעה"])
        rph_chart = rph_chart.sort_values("מכר לשעה", ascending=True)
        if not rph_chart.empty:
            st.bar_chart(
                rph_chart.set_index("סניף")["מכר לשעה"],
                horizontal=True,
                use_container_width=True,
            )
        else:
            st.info("אין נתוני שעות פעילות לחישוב מכר לשעה.")

    with tab3:
        portions_chart = merged[["סניף", "מנות בפיתה", "ארוחות בפיתה"]].copy()
        portions_chart = portions_chart.dropna(subset=["מנות בפיתה"])
        portions_chart = portions_chart.sort_values("מנות בפיתה", ascending=True)
        st.bar_chart(
            portions_chart.set_index("סניף")[["מנות בפיתה", "ארוחות בפיתה"]],
            horizontal=True,
            use_container_width=True,
        )

    with tab4:
        meal_chart = ranking_df[["סניף", "מנות בפיתה", "ארוחות בפיתה"]].copy()
        meal_chart = meal_chart[(meal_chart["מנות בפיתה"].notna()) & (meal_chart["מנות בפיתה"] > 0)]
        meal_chart["אחוז ארוחות"] = meal_chart["ארוחות בפיתה"] / meal_chart["מנות בפיתה"]
        meal_chart = meal_chart.sort_values("אחוז ארוחות", ascending=True)
        if not meal_chart.empty:
            st.bar_chart(
                meal_chart.set_index("סניף")["אחוז ארוחות"],
                horizontal=True,
                use_container_width=True,
            )
        else:
            st.info("אין נתוני מנות לחישוב אחוז ארוחות.")

    # ── Data Table ──
    st.markdown("---")
    st.subheader("📋 טבלת נתונים")

    # Format for display
    display_df = merged.copy()
    display_df['מכר כולל מע"מ'] = display_df['מכר כולל מע"מ'].apply(
        lambda x: f"₪{x:,.0f}" if pd.notna(x) else ""
    )
    display_df["ממוצע עסקאות"] = display_df["ממוצע עסקאות"].apply(
        lambda x: f"{x:.2f}" if pd.notna(x) else ""
    )
    display_df["מס' עסקאות"] = display_df["מס' עסקאות"].apply(
        lambda x: f"{int(x)}" if pd.notna(x) else ""
    )
    display_df["עסקה ראשונה"] = display_df["עסקה ראשונה"].apply(
        lambda x: x.strftime("%H:%M") if isinstance(x, datetime.time) else (str(x) if pd.notna(x) else "")
    )
    display_df["עסקה אחרונה"] = display_df["עסקה אחרונה"].apply(
        lambda x: x.strftime("%H:%M") if isinstance(x, datetime.time) else (str(x) if pd.notna(x) else "")
    )
    display_df["מנות בפיתה"] = display_df["מנות בפיתה"].apply(
        lambda x: f"{int(x)}" if pd.notna(x) else ""
    )
    display_df["ארוחות בפיתה"] = display_df["ארוחות בפיתה"].apply(
        lambda x: f"{int(x)}" if pd.notna(x) else ""
    )
    display_df["אחוז ארוחות מתוך מנות"] = display_df["אחוז ארוחות מתוך מנות"].apply(
        lambda x: f"{x:.1%}" if pd.notna(x) and x > 0 else "0%"
    )

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
    )

    # ── Download Button (both in sidebar and main area) ──
    st.markdown("---")

    excel_bytes = generate_excel(merged, report_date)

    date_str = report_date.strftime("%d.%m.%y")
    file_name = f"דוח_יומי_{date_str}.xlsx"

    # Sidebar download — always visible
    with sidebar_download_placeholder:
        st.divider()
        st.download_button(
            label=f"⬇️ הורד דוח יומי - {report_date.strftime('%d/%m/%Y')}",
            data=excel_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sidebar_download",
        )

    # Main area download too (for discoverability)
    st.download_button(
        label=f"⬇️ הורד דוח יומי - {report_date.strftime('%d/%m/%Y')}",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="main_download",
    )

    # ── Raw data expander ──
    with st.expander("🔧 נתונים גולמיים (Debug)", expanded=False):
        if csv_data is not None and not csv_data.empty:
            st.write("**CSV מכר:**")
            st.dataframe(csv_data, hide_index=True)
        if avg_trans_data is not None and not avg_trans_data.empty:
            st.write("**ממוצע עסקה:**")
            st.dataframe(avg_trans_data, hide_index=True)
        if portions_data is not None and not portions_data.empty:
            st.write("**מנות בפיתה:**")
            st.dataframe(portions_data, hide_index=True)
        if hourly_data is not None and not hourly_data.empty:
            st.write("**עסקאות לפי שעה:**")
            st.dataframe(hourly_data, hide_index=True)
        for i, ps in enumerate(paz_sales_list):
            st.write(f"**PDF מכירות פז #{i+1}:**")
            st.dataframe(ps, hide_index=True)
        for i, pp in enumerate(paz_portions_list):
            st.write(f"**PDF מנות פז #{i+1}:**")
            st.dataframe(pp, hide_index=True)


if __name__ == "__main__":
    main()
