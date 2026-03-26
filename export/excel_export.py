"""Generate a professionally formatted 2-sheet Excel report."""

import io
import datetime
import pandas as pd

from config import MASTER_COLUMNS, PAZ_BRANCHES


def generate_excel(df: pd.DataFrame, report_date: datetime.date) -> bytes:
    """Generate a professionally formatted Excel report.

    Layout (compact, no wasted rows):
      Row 1 : Title bar — date + report name, merged across all columns
      Row 2 : Column headers (blue background, white bold text)
      Row 3…: Data rows (alternating zebra stripes)
      Last  : Totals row (bold, light-blue background)

    Uses only in-memory io.BytesIO — works identically on Windows and macOS.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    # Sheet name: DD.M format (like "15.2")
    ws.title = f"{report_date.day}.{report_date.month}"

    # ── Styles ──
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(size=11)
    branch_font = Font(size=11, bold=True)
    totals_font = Font(bold=True, size=11, color="1F3864")
    totals_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    stripe_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    num_cols = len(MASTER_COLUMNS)

    # ── Row 1: Title bar (merged) ──
    date_str = report_date.strftime("%d.%m.%Y")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"דוח מכירות יומי — {date_str}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 32

    for ci in range(2, num_cols + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = title_fill
        c.border = thin_border

    # ── Row 2: Column headers ──
    ws.row_dimensions[2].height = 28
    for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # ── Prepare export data ──
    export_df = df.copy()

    for col in ["עסקה ראשונה", "עסקה אחרונה"]:
        export_df[col] = export_df[col].apply(
            lambda x: x.strftime("%H:%M") if isinstance(x, datetime.time) else (
                str(x) if pd.notna(x) else ""
            )
        )

    # ── Rows 3…N: Data ──
    data_start_row = 3
    for ri, (_, row) in enumerate(export_df.iterrows()):
        excel_row = data_start_row + ri
        is_stripe = ri % 2 == 1

        for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
            val = row.get(col_name)
            cell = ws.cell(row=excel_row, column=ci)
            cell.border = thin_border

            if pd.isna(val):
                cell.value = None
            else:
                cell.value = val

            if col_name == "סניף":
                cell.font = branch_font
                cell.alignment = right_align
            elif col_name == 'מכר כולל מע"מ':
                cell.number_format = "#,##0"
                cell.font = data_font
                cell.alignment = center
            elif col_name == "ממוצע עסקאות":
                cell.number_format = "#,##0.00"
                cell.font = data_font
                cell.alignment = center
            elif col_name == "אחוז ארוחות מתוך מנות":
                cell.number_format = "0.00%"
                cell.font = data_font
                cell.alignment = center
            elif col_name in ("מנות בפיתה", "ארוחות בפיתה", "מס' עסקאות"):
                cell.number_format = "#,##0"
                cell.font = data_font
                cell.alignment = center
            else:
                cell.font = data_font
                cell.alignment = center

            if is_stripe:
                cell.fill = stripe_fill

    # ── Totals row ──
    totals_row_idx = data_start_row + len(export_df)
    totals_values = {
        "סניף": 'סה"כ',
        'מכר כולל מע"מ': df['מכר כולל מע"מ'].sum(),
        "ממוצע עסקאות": None,
        "מס' עסקאות": df["מס' עסקאות"].sum() if "מס' עסקאות" in df.columns else None,
        "עסקה ראשונה": None,
        "עסקה אחרונה": None,
        "מנות בפיתה": df["מנות בפיתה"].sum(),
        "ארוחות בפיתה": df["ארוחות בפיתה"].sum(),
    }
    total_p = totals_values["מנות בפיתה"] or 0
    total_m = totals_values["ארוחות בפיתה"] or 0
    totals_values["אחוז ארוחות מתוך מנות"] = total_m / total_p if total_p > 0 else 0

    for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=totals_row_idx, column=ci, value=totals_values.get(col_name))
        cell.font = totals_font
        cell.fill = totals_fill
        cell.alignment = center
        cell.border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="medium", color="4472C4"),
            bottom=Side(style="medium", color="4472C4"),
        )
        if col_name == 'מכר כולל מע"מ':
            cell.number_format = "#,##0"
        elif col_name == "אחוז ארוחות מתוך מנות":
            cell.number_format = "0.00%"
        elif col_name in ("מנות בפיתה", "ארוחות בפיתה", "מס' עסקאות"):
            cell.number_format = "#,##0"

    # ── Column widths ──
    from openpyxl.utils import get_column_letter

    col_widths = {
        "סניף": 16,
        'מכר כולל מע"מ': 16,
        "ממוצע עסקאות": 15,
        "מס' עסקאות": 13,
        "עסקה ראשונה": 15,
        "עסקה אחרונה": 15,
        "מנות בפיתה": 14,
        "ארוחות בפיתה": 15,
        "אחוז ארוחות מתוך מנות": 22,
    }
    for ci, col_name in enumerate(MASTER_COLUMNS, start=1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A3"
    ws.sheet_view.rightToLeft = True
    ws.print_title_rows = "1:2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ══════════════════════════════════════════════
    # Sheet 2: "ניתוח" (Analysis)
    # ══════════════════════════════════════════════
    _build_analysis_sheet(wb, df, report_date, ws.title)

    wb.save(output)
    return output.getvalue()


def _build_analysis_sheet(wb, df: pd.DataFrame, report_date: datetime.date, data_sheet_name: str):
    """Build a second sheet with derived analytics, rankings, and charts.

    Contents:
      - Section A: Branch ranking table (sorted by revenue) with extra
        derived columns: revenue %, operating hours, revenue/hour
      - Section B: Paz vs Non-Paz group summary
      - Section C: 4 charts (revenue, pie, efficiency, meal %)
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList

    ws2 = wb.create_sheet(title="ניתוח")
    ws2.sheet_view.rightToLeft = True

    # ── Shared styles ──
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(size=10)
    bold_font = Font(size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    stripe = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    subtotal_font = Font(bold=True, size=10, color="1F3864")

    num_branches = len(df)
    total_rev = df['מכר כולל מע"מ'].sum()

    # ── Prepare analysis DataFrame (sorted by revenue descending) ──
    adf = df.copy().sort_values('מכר כולל מע"מ', ascending=False).reset_index(drop=True)

    adf["% מהמכר"] = adf['מכר כולל מע"מ'] / total_rev if total_rev > 0 else 0

    def _calc_hours(row):
        f = row.get("עסקה ראשונה")
        l = row.get("עסקה אחרונה")
        if isinstance(f, datetime.time) and isinstance(l, datetime.time):
            f_min = f.hour * 60 + f.minute
            l_min = l.hour * 60 + l.minute
            diff = l_min - f_min
            return round(diff / 60, 1) if diff > 0 else None
        return None
    adf["שעות פעילות"] = adf.apply(_calc_hours, axis=1)

    adf["מכר לשעה"] = adf.apply(
        lambda r: round(r['מכר כולל מע"מ'] / r["שעות פעילות"])
        if r["שעות פעילות"] and r["שעות פעילות"] > 0 else None,
        axis=1,
    )

    adf["_paz"] = adf["סניף"].isin(PAZ_BRANCHES)

    # ────────────────────────────────
    # Row 1: Title
    # ────────────────────────────────
    analysis_cols = [
        "דירוג", "סניף", 'מכר כולל מע"מ', "% מהמכר",
        "מס' עסקאות", "שעות פעילות", "מכר לשעה",
        "מנות בפיתה", "ארוחות בפיתה",
    ]
    ncols = len(analysis_cols)

    date_str = report_date.strftime("%d.%m.%Y")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws2.cell(row=1, column=1, value=f"ניתוח יומי — {date_str}")
    tc.font = title_font
    tc.fill = title_fill
    tc.alignment = center
    tc.border = thin_border
    ws2.row_dimensions[1].height = 32
    for ci in range(2, ncols + 1):
        c = ws2.cell(row=1, column=ci)
        c.fill = title_fill
        c.border = thin_border

    # ────────────────────────────────
    # Section A: Ranked Branch Table
    # ────────────────────────────────
    sec_a_row = 3
    ws2.cell(row=sec_a_row, column=1, value="דירוג סניפים לפי מכר").font = section_font

    hdr_row = sec_a_row + 1
    ws2.row_dimensions[hdr_row].height = 26
    for ci, col_name in enumerate(analysis_cols, start=1):
        cell = ws2.cell(row=hdr_row, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    data_start = hdr_row + 1
    for ri, (_, row) in enumerate(adf.iterrows()):
        er = data_start + ri
        rank = ri + 1
        is_stripe = ri % 2 == 1

        values = [
            rank,
            row["סניף"],
            row['מכר כולל מע"מ'],
            row["% מהמכר"],
            row.get("מס' עסקאות"),
            row.get("שעות פעילות"),
            row.get("מכר לשעה"),
            row.get("מנות בפיתה"),
            row.get("ארוחות בפיתה"),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws2.cell(row=er, column=ci)
            cell.value = val if pd.notna(val) else None
            cell.border = thin_border
            cell.alignment = center
            cell.font = bold_font if ci == 2 else data_font

            col_name = analysis_cols[ci - 1]
            if col_name == 'מכר כולל מע"מ':
                cell.number_format = "#,##0"
            elif col_name == "% מהמכר":
                cell.number_format = "0.0%"
            elif col_name == "מכר לשעה":
                cell.number_format = "#,##0"
            elif col_name in ("מנות בפיתה", "ארוחות בפיתה", "מס' עסקאות"):
                cell.number_format = "#,##0"
            elif col_name == "שעות פעילות":
                cell.number_format = "0.0"

            if is_stripe:
                cell.fill = stripe

        if rank <= 3:
            for ci in range(1, ncols + 1):
                ws2.cell(row=er, column=ci).fill = gold_fill
        elif rank > num_branches - 3:
            for ci in range(1, ncols + 1):
                ws2.cell(row=er, column=ci).fill = red_fill

    # ────────────────────────────────
    # Section B: Paz vs Non-Paz Summary
    # ────────────────────────────────
    sec_b_row = data_start + num_branches + 2
    ws2.cell(row=sec_b_row, column=1, value="סיכום: רשת מול פז").font = section_font

    summary_cols = ["קבוצה", "סניפים", 'מכר כולל מע"מ', "% מהמכר", "מס' עסקאות", "ממוצע מכר לסניף"]
    shdr = sec_b_row + 1
    ws2.row_dimensions[shdr].height = 26
    for ci, cn in enumerate(summary_cols, start=1):
        cell = ws2.cell(row=shdr, column=ci, value=cn)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    paz_df = adf[adf["_paz"]]
    non_paz_df = adf[~adf["_paz"]]

    groups = [
        ("סניפי רשת", non_paz_df),
        ("סניפי פז", paz_df),
        ('סה"כ', adf),
    ]
    for gi, (label, gdf) in enumerate(groups):
        r = shdr + 1 + gi
        g_rev = gdf['מכר כולל מע"מ'].sum()
        g_count = len(gdf)
        g_trans = gdf["מס' עסקאות"].sum() if "מס' עסקאות" in gdf.columns else 0
        g_pct = g_rev / total_rev if total_rev > 0 else 0
        g_avg = g_rev / g_count if g_count > 0 else 0

        vals = [label, g_count, g_rev, g_pct, g_trans, g_avg]
        is_total = (gi == 2)
        for ci, val in enumerate(vals, start=1):
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = center
            cell.font = subtotal_font if is_total else bold_font
            if is_total:
                cell.fill = subtotal_fill
            elif gi == 0:
                cell.fill = green_fill

            cn = summary_cols[ci - 1]
            if cn == 'מכר כולל מע"מ':
                cell.number_format = "#,##0"
            elif cn == "% מהמכר":
                cell.number_format = "0.0%"
            elif cn == "ממוצע מכר לסניף":
                cell.number_format = "#,##0"
            elif cn == "מס' עסקאות":
                cell.number_format = "#,##0"

    # ────────────────────────────────
    # Section C: Charts
    # ────────────────────────────────
    # Charts are laid out in a single vertical column with generous
    # spacing so they never overlap — critical for Google Sheets compat.
    CHART_H   = 14          # cm – height for full-width bar charts
    CHART_W   = 24          # cm – width for full-width bar charts
    SMALL_H   = 12          # cm – height for smaller charts (pie)
    SMALL_W   = 18          # cm – width for smaller charts (pie)
    CHART_GAP = 34          # rows between chart anchors

    # Helper data area starts far to the right (col 14+)
    HELPER_COL = 14

    chart_row = shdr + 5

    # ── Chart 1: Revenue by branch (horizontal bar) ──
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = "מכר לפי סניף"
    chart1.y_axis.title = None
    chart1.x_axis.title = None
    chart1.style = 26
    chart1.width = CHART_W
    chart1.height = CHART_H
    chart1.legend = None

    data_ref = Reference(ws2, min_col=3, min_row=hdr_row,
                         max_row=data_start + num_branches - 1, max_col=3)
    cats_ref = Reference(ws2, min_col=2, min_row=data_start,
                         max_row=data_start + num_branches - 1)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.solidFill = "4472C4"
    ws2.add_chart(chart1, f"A{chart_row}")

    # ── Chart 2: Paz vs Non-Paz pie chart ──
    chart_row += CHART_GAP

    pie_data_row = chart_row
    ws2.cell(row=pie_data_row, column=HELPER_COL, value="קבוצה")
    ws2.cell(row=pie_data_row, column=HELPER_COL + 1, value="מכר")
    paz_rev = adf.loc[adf["_paz"], 'מכר כולל מע"מ'].sum()
    non_paz_rev = adf.loc[~adf["_paz"], 'מכר כולל מע"מ'].sum()
    ws2.cell(row=pie_data_row + 1, column=HELPER_COL, value="סניפי רשת")
    ws2.cell(row=pie_data_row + 1, column=HELPER_COL + 1, value=non_paz_rev)
    ws2.cell(row=pie_data_row + 2, column=HELPER_COL, value="סניפי פז")
    ws2.cell(row=pie_data_row + 2, column=HELPER_COL + 1, value=paz_rev)

    pie = PieChart()
    pie.title = "פילוח מכר: רשת מול פז"
    pie.style = 26
    pie.width = SMALL_W
    pie.height = SMALL_H

    pie_data = Reference(ws2, min_col=HELPER_COL + 1, min_row=pie_data_row,
                         max_row=pie_data_row + 2)
    pie_cats = Reference(ws2, min_col=HELPER_COL, min_row=pie_data_row + 1,
                         max_row=pie_data_row + 2)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)

    slice_chain = DataPoint(idx=0)
    slice_chain.graphicalProperties.solidFill = "4472C4"
    slice_paz = DataPoint(idx=1)
    slice_paz.graphicalProperties.solidFill = "F2994A"
    pie.series[0].data_points = [slice_chain, slice_paz]

    pie.series[0].dLbls = DataLabelList()
    pie.series[0].dLbls.showPercent = True
    pie.series[0].dLbls.showCatName = True
    pie.series[0].dLbls.showVal = False

    ws2.add_chart(pie, f"A{chart_row}")

    # ── Chart 3: Revenue per hour (efficiency) ──
    chart_row += CHART_GAP

    rph_header_row = chart_row
    ws2.cell(row=rph_header_row, column=HELPER_COL, value="סניף")
    ws2.cell(row=rph_header_row, column=HELPER_COL + 1, value="מכר לשעה")
    rph_df = adf[adf["מכר לשעה"].notna()].sort_values("מכר לשעה", ascending=False)
    for ri, (_, row) in enumerate(rph_df.iterrows()):
        ws2.cell(row=rph_header_row + 1 + ri, column=HELPER_COL, value=row["סניף"])
        ws2.cell(row=rph_header_row + 1 + ri, column=HELPER_COL + 1, value=row["מכר לשעה"])
    rph_count = len(rph_df)

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "מכר לשעת פעילות"
    chart3.y_axis.title = None
    chart3.x_axis.title = None
    chart3.style = 26
    chart3.width = CHART_W
    chart3.height = CHART_H
    chart3.legend = None

    rph_data_ref = Reference(ws2, min_col=HELPER_COL + 1, min_row=rph_header_row,
                             max_row=rph_header_row + rph_count)
    rph_cats_ref = Reference(ws2, min_col=HELPER_COL, min_row=rph_header_row + 1,
                             max_row=rph_header_row + rph_count)
    chart3.add_data(rph_data_ref, titles_from_data=True)
    chart3.set_categories(rph_cats_ref)
    chart3.series[0].graphicalProperties.solidFill = "11998E"
    ws2.add_chart(chart3, f"A{chart_row}")

    # ── Chart 4: Meal % of portions ──
    chart_row += CHART_GAP

    meal_header_row = chart_row
    ws2.cell(row=meal_header_row, column=HELPER_COL, value="סניף")
    ws2.cell(row=meal_header_row, column=HELPER_COL + 1, value="% ארוחות")
    meal_df = adf[
        (adf["מנות בפיתה"].notna()) & (adf["מנות בפיתה"] > 0) &
        (adf["ארוחות בפיתה"].notna())
    ].copy()
    meal_df["_meal_pct"] = meal_df["ארוחות בפיתה"] / meal_df["מנות בפיתה"]
    meal_df = meal_df.sort_values("_meal_pct", ascending=False)
    for ri, (_, row) in enumerate(meal_df.iterrows()):
        ws2.cell(row=meal_header_row + 1 + ri, column=HELPER_COL, value=row["סניף"])
        ws2.cell(row=meal_header_row + 1 + ri, column=HELPER_COL + 1, value=row["_meal_pct"])
        ws2.cell(row=meal_header_row + 1 + ri, column=HELPER_COL + 1).number_format = "0.0%"
    meal_count = len(meal_df)

    chart4 = BarChart()
    chart4.type = "bar"
    chart4.title = "אחוז ארוחות מתוך מנות בפיתה"
    chart4.y_axis.title = None
    chart4.x_axis.title = None
    chart4.style = 26
    chart4.width = CHART_W
    chart4.height = CHART_H
    chart4.legend = None

    meal_data_ref = Reference(ws2, min_col=HELPER_COL + 1, min_row=meal_header_row,
                              max_row=meal_header_row + meal_count)
    meal_cats_ref = Reference(ws2, min_col=HELPER_COL, min_row=meal_header_row + 1,
                              max_row=meal_header_row + meal_count)
    chart4.add_data(meal_data_ref, titles_from_data=True)
    chart4.set_categories(meal_cats_ref)
    chart4.series[0].graphicalProperties.solidFill = "F2C94C"
    ws2.add_chart(chart4, f"A{chart_row}")

    # ── Column widths ──
    widths = [8, 16, 16, 12, 13, 13, 14, 14, 14]
    for ci, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = "A5"
